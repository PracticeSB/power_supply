import pyvisa
import time
import datetime
import openpyxl
import keyboard
import numpy as np

#=================================================================================
# Making a worksheet
#=================================================================================
wb = openpyxl.Workbook()
ws = wb.create_sheet("TEST")  # Name of the sheet
ws.cell(row=1, column=1, value='Time[Sec]')
ws.cell(row=1, column=2, value='Input_voltage[A]')
ws.cell(row=1, column=3, value='Resistance[Ohms]')
#=================================================================================

rm = 0
my_PS = 0
Total_time = 0
Loop_count = 1
Set_point = 15
i_control = 0
pi = 0


d = datetime.datetime.now()

def datalogging(x,y,data):
    ws.cell(row=x, column=y, value=data)

def dataoutput(finish_time):
    wb.save(r'D:\Dropbox\노승범\실험\TOF\Laser\20220728\220%\Fabrication_170g\20221214\90g_retry\Practice_'
            +finish_time.strftime('%Y%m%d%H%M')+'.xlsx')
    print('Data is saved!')

def Input_power(power_cm,sample_length):
    print('Input power/cm: ', power_cm)
    print('Sample length: ', sample_length)
    return power_cm*sample_length

def map(x,input_min,input_max,output_min,output_max):
    return (x-input_min)*(output_max-output_min)/(input_max-input_min)+output_min

def KEI2231_Connect(rsrcString, getIdStr, timeout, doRst):
    my_PS = rm.open_resource(rsrcString, baud_rate = 9600, data_bits = 8)	#opens desired resource and sets it variable my_instrument
    my_PS.write_termination = '\n'
    my_PS.read_termination = '\n'
    my_PS.send_end = True
    my_PS.StopBits = 1
    if getIdStr == 1:
        print(my_PS.query("*IDN?"))
        my_PS.write('*RST')
        print('Initialization is completed')
        print('=========================================')
    my_PS.write('SYST:REM')
    my_PS.timeout = timeout
    if doRst == 1:
        return my_PS

def KEI2231A_Disconnect():
    my_PS.write('SYST:LOC')
    my_PS.close
    return

def KEI2231A_SelectChannel(myChan):
    my_PS.write("INST:NSEL %d" % myChan)
    return

def KEI2231A_SetVoltage(myV,myI):
    my_PS.write('APPLy CH1,%0.4f,%0.4f' %(myV,myI)) #마지막은 current 범위
    return myV

def KEI2231A_OutputState(myState):
    if myState == 0:
        my_PS.write("OUTP 0")
    else:
        my_PS.write("OUTP 1")
    return
#================================================================================
# MAIN CODE GOES HERE
#================================================================================

rm = pyvisa.ResourceManager()   # Opens the resource manager and sets it to variable rm
my_PS = KEI2231_Connect("ASRL5::INSTR", 1, 20000, 1) #VISA communication
KEI2231A_SelectChannel(1)
KEI2231A_OutputState(1)     # 1일 경우에는 ON
#=================================================================================
# Getting Initial resistance value
#=================================================================================

Initial_voltage = 3
KEI2231A_SetVoltage(Initial_voltage,1)
time.sleep(1)
my_PS.write('FETC:CURR?')
Current = my_PS.read()
print('Initial current: ', Current)
my_PS.write('FETC:CURR?')
Current = my_PS.read()
Initial_Resistance = Initial_voltage/float(Current)
print('Initial resistance: ', Initial_Resistance)
Target_power = Input_power(0.15,5)
Target_voltage = np.sqrt(Target_power*Initial_Resistance)
print('Target_voltage:',Target_voltage)
print('=========================================')

while True:
    global t1
    global t2
    global Kp
    global Ki
    Ki = 0.1
    Kp = 1.5
    if keyboard.is_pressed("a"):
        break
    t1 = time.time()    # Capture start time....a
    Target_voltage = Target_voltage
    try:
        KEI2231A_SetVoltage(Target_voltage,1)
        my_PS.write('FETC:CURR?')
        Current = my_PS.read()
        Resistance = Target_voltage/float(Current)
        if Resistance > Initial_Resistance*2:
            #print('Error value')
            my_PS.write('FETC:CURR?')
            time.sleep(0.1)
            Current = my_PS.read()
            Resistance = Target_voltage/float(Current)
        else:
            pass
    except ZeroDivisionError:
        pass
    #print('Current value: ',"%0.4f" %float(Current))
    #print('Resistance value:', "%0.4f" %Resistance)
    t2 = time.time()    # Capture stop time...
    #print("Loop time:","{0:.4f}s".format((t2-t1)),",",Loop_count)
    Delta_r_percentage = -((Resistance-float(Initial_Resistance))/float(Initial_Resistance))*100
    error = Set_point-Delta_r_percentage
    print(error)
    p_control = Kp*error
    i_control = i_control+Ki*error*(t2 - t1)
    pi = p_control+i_control #Output data
    Target_voltage = map(pi,0,100,0,10) #pi control의 값을 voltage로 치환하여서 대입함.
    print("Target_voltage: ", Target_voltage)
    Total_time = (t2 - t1) + Total_time
    Loop_count = Loop_count + 1
    datalogging(Loop_count+2,1,Total_time)
    datalogging(Loop_count+2,2,Target_voltage)
    datalogging(Loop_count+2,3,Resistance)

dataoutput(d)
KEI2231A_OutputState(0)
KEI2231A_Disconnect()

rm.close



